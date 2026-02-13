"""
Generate test XLSX files with varying similarity levels for testing.
Run this once, then delete it.
"""
from openpyxl import Workbook
import os

output_folder = "xlsx_files"
os.makedirs(output_folder, exist_ok=True)


def save_wb(wb, filename):
    wb.save(os.path.join(output_folder, filename))
    print(f"Created {filename}")


# ============================================================
# FILE 1: Original - Q1 Budget Report
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "Q1 Budget"
ws.append(["Department", "Budget (EUR)", "Spent (EUR)", "Remaining (EUR)", "Status"])
ws.append(["Engineering", 120000, 95000, 25000, "On Track"])
ws.append(["Marketing", 80000, 72000, 8000, "At Risk"])
ws.append(["Sales", 65000, 58000, 7000, "At Risk"])
ws.append(["HR", 45000, 30000, 15000, "On Track"])
ws.append(["Finance", 35000, 28000, 7000, "On Track"])
ws.append(["Total", 345000, 283000, 62000, "On Track"])

ws2 = wb.create_sheet("Notes")
ws2.append(["Q1 2025 Budget Summary"])
ws2.append(["Marketing overspend due to trade show in March"])
ws2.append(["Sales team travel costs higher than expected"])
ws2.append(["Engineering saved costs by migrating to cloud infrastructure"])
ws2.append(["Overall budget utilization at 82 percent which is within target"])
save_wb(wb, "q1_budget_report.xlsx")

# ============================================================
# FILE 2: EXACT DUPLICATE (~1.0 similarity)
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "Q1 Budget"
ws.append(["Department", "Budget (EUR)", "Spent (EUR)", "Remaining (EUR)", "Status"])
ws.append(["Engineering", 120000, 95000, 25000, "On Track"])
ws.append(["Marketing", 80000, 72000, 8000, "At Risk"])
ws.append(["Sales", 65000, 58000, 7000, "At Risk"])
ws.append(["HR", 45000, 30000, 15000, "On Track"])
ws.append(["Finance", 35000, 28000, 7000, "On Track"])
ws.append(["Total", 345000, 283000, 62000, "On Track"])

ws2 = wb.create_sheet("Notes")
ws2.append(["Q1 2025 Budget Summary"])
ws2.append(["Marketing overspend due to trade show in March"])
ws2.append(["Sales team travel costs higher than expected"])
ws2.append(["Engineering saved costs by migrating to cloud infrastructure"])
ws2.append(["Overall budget utilization at 82 percent which is within target"])
save_wb(wb, "q1_budget_report_copy.xlsx")

# ============================================================
# FILE 3: UPDATED VERSION (~0.85-0.95 similarity)
# Same structure, slightly different numbers
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "Q1 Budget"
ws.append(["Department", "Budget (EUR)", "Spent (EUR)", "Remaining (EUR)", "Status"])
ws.append(["Engineering", 120000, 98000, 22000, "On Track"])
ws.append(["Marketing", 80000, 78000, 2000, "Over Budget"])
ws.append(["Sales", 65000, 60000, 5000, "At Risk"])
ws.append(["HR", 45000, 32000, 13000, "On Track"])
ws.append(["Finance", 35000, 29000, 6000, "On Track"])
ws.append(["Total", 345000, 297000, 48000, "At Risk"])

ws2 = wb.create_sheet("Notes")
ws2.append(["Q1 2025 Budget Summary - REVISED"])
ws2.append(["Marketing exceeded budget due to additional trade show in March and April campaign"])
ws2.append(["Sales team travel costs significantly higher than expected"])
ws2.append(["Engineering saved costs by migrating to cloud infrastructure"])
ws2.append(["Overall budget utilization at 86 percent which is slightly above target"])
save_wb(wb, "q1_budget_report_v2.xlsx")

# ============================================================
# FILE 4: DIFFERENT QUARTER SAME FORMAT (~0.7-0.8 similarity)
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "Q2 Budget"
ws.append(["Department", "Budget (EUR)", "Spent (EUR)", "Remaining (EUR)", "Status"])
ws.append(["Engineering", 130000, 65000, 65000, "On Track"])
ws.append(["Marketing", 90000, 45000, 45000, "On Track"])
ws.append(["Sales", 70000, 38000, 32000, "On Track"])
ws.append(["HR", 50000, 22000, 28000, "On Track"])
ws.append(["Finance", 40000, 18000, 22000, "On Track"])
ws.append(["Total", 380000, 188000, 192000, "On Track"])

ws2 = wb.create_sheet("Notes")
ws2.append(["Q2 2025 Budget Summary"])
ws2.append(["All departments currently under budget at midpoint"])
ws2.append(["Marketing budget increased to support France expansion"])
ws2.append(["New Paris office costs allocated to Engineering and Sales"])
ws2.append(["Overall budget utilization at 49 percent which is on target for mid-quarter"])
save_wb(wb, "q2_budget_report.xlsx")

# ============================================================
# FILE 5: DIFFERENT TOPIC (~0.3-0.5 similarity)
# Employee contact list - completely different
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "Employees"
ws.append(["Name", "Department", "Email", "Phone", "Office", "Start Date"])
ws.append(["Jan Peeters", "Engineering", "jan.peeters@company.com", "+32 471 123 456", "Brussels", "2020-03-15"])
ws.append(["Sophie De Vries", "Marketing", "sophie.devries@company.com", "+31 612 345 678", "Amsterdam", "2021-06-01"])
ws.append(["Thomas Mueller", "Sales", "thomas.mueller@company.com", "+49 170 987 654", "Munich", "2019-11-20"])
ws.append(["Marie Dubois", "HR", "marie.dubois@company.com", "+32 473 456 789", "Brussels", "2022-01-10"])
ws.append(["Lars Andersen", "Finance", "lars.andersen@company.com", "+45 20 123 456", "Copenhagen", "2023-04-01"])

ws2 = wb.create_sheet("Notes")
ws2.append(["Employee Directory - Last Updated January 2025"])
ws2.append(["Contact IT helpdesk for access issues"])
ws2.append(["Report changes to HR department"])
save_wb(wb, "employee_directory.xlsx")

print(f"\nDone! Created 5 test XLSX files in '{output_folder}/'")
print("\nExpected similarity levels:")
print("  q1_budget_report  vs  q1_budget_report_copy  → ~1.0  (exact duplicate)")
print("  q1_budget_report  vs  q1_budget_report_v2    → ~0.85-0.95 (revised numbers)")
print("  q1_budget_report  vs  q2_budget_report       → ~0.7-0.8  (different quarter)")
print("  q1_budget_report  vs  employee_directory      → ~0.3-0.5  (different topic)")
